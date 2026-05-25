"""
=============================================================
  AGENT PROSPECTION GSA PRADO — v3
  Point de départ : prospection_v2.xlsx (14 colonnes)
  Ce script ajoute : Odoo CRM + Zeliq + Création leads
=============================================================

INSTALLATION (une seule fois) :
  pip install requests openpyxl

USAGE :
  1. Placez ce script dans le même dossier que prospection_v2.xlsx
  2. Lancez : python odoo_prospection_v3.py
  3. Suivez le menu interactif

COLONNES AJOUTÉES PAR CE SCRIPT :
  O (15) — Statut Odoo CRM
  P (16) — URL LinkedIn à coller manuellement
  Q (17) — Email enrichi (Zeliq)
  R (18) — Téléphone enrichi (Zeliq)
  S (19) — Statut enrichissement Zeliq
=============================================================
"""

import xmlrpc.client
import re, os, sys, time
from datetime import datetime

try:
    import requests
except ImportError:
    print("❌ Module 'requests' manquant.\n   Lancez : pip install requests openpyxl")
    sys.exit(1)

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("❌ Module 'openpyxl' manquant.\n   Lancez : pip install openpyxl")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
#  CONFIG
# ══════════════════════════════════════════════════════════════════════════════

ODOO_URL     = "https://gsa-prado.odoo.com"
ODOO_DB      = "unikerp-gsaprado-prod-13772120"
ODOO_USER    = "georges-eric.michel@gsaprado.fr"
ODOO_API_KEY = "b4b43e5d24ac0631710e427dbfacc7951bcc7095"

ZELIQ_API_KEY = "sk-c4ce11eee8cfc5bf38c6256bbb27dd0dc85e65c68a52f740"
ZELIQ_BASE    = "https://api.zeliq.com/api"
ZELIQ_HDR     = {"x-api-key": ZELIQ_API_KEY, "Content-Type": "application/json"}

# Fichiers
INPUT_FILE  = "prospection_v2.xlsx"   # votre fichier enrichi généré par Claude
OUTPUT_FILE = "prospection_v3_final.xlsx"

# ── Colonnes existantes dans prospection_v2.xlsx ──────────────────────────────
COL_ENTREPRISE = 1   # A
COL_SECTEUR    = 2   # B
COL_VILLE      = 3   # C
COL_DG         = 8   # H — 🥇 Décideur final
COL_DAF        = 9   # I — 🥈 Décideur opérationnel
COL_SIGNAL     = 11  # K
COL_ACCROCHE   = 12  # L

# ── Nouvelles colonnes ajoutées par ce script ─────────────────────────────────
COL_ODOO      = 15   # O — Statut Odoo
COL_LINKEDIN  = 16   # P — URL LinkedIn (saisie manuelle)
COL_EMAIL     = 17   # Q — Email Zeliq
COL_PHONE     = 18   # R — Téléphone Zeliq
COL_ZELIQ_ST  = 19   # S — Statut Zeliq

FIRST_ROW = 2        # première ligne de données (après l'en-tête)

# ── Couleurs ──────────────────────────────────────────────────────────────────
BG = {
    "header":    "1A3A5C",
    "green":     "D5F5E3",
    "blue":      "D6EAF8",
    "red":       "FADBD8",
    "gray":      "F2F3F4",
    "yellow":    "FEF9E7",
    "purple":    "EDE7F6",
    "linkedin":  "E8F4FD",
    "white":     "FFFFFF",
}

def fill(c):    return PatternFill("solid", fgColor=c)
def border():
    s = Side(style='thin', color="BDC3C7")
    return Border(left=s, right=s, top=s, bottom=s)
def align(h='left'):
    return Alignment(horizontal=h, vertical='center', wrap_text=True)

def write(ws, row, col, value, bg="white", bold=False, color="1A1A2E", center=False):
    c = ws.cell(row=row, column=col, value=str(value) if value else '')
    c.font      = Font(name='Arial', size=9, bold=bold, color=color)
    c.fill      = fill(BG.get(bg, bg))
    c.alignment = align('center' if center else 'left')
    c.border    = border()
    return c


# ══════════════════════════════════════════════════════════════════════════════
#  INITIALISATION DU FICHIER
# ══════════════════════════════════════════════════════════════════════════════

def init_file():
    """Charge le fichier source et ajoute les en-têtes des nouvelles colonnes."""
    if not os.path.exists(INPUT_FILE):
        print(f"\n❌ Fichier introuvable : '{INPUT_FILE}'")
        print(f"   Assurez-vous que '{INPUT_FILE}' est dans le même dossier que ce script.")
        sys.exit(1)

    # Copier dans le fichier de sortie si pas encore fait
    if not os.path.exists(OUTPUT_FILE):
        import shutil
        shutil.copy(INPUT_FILE, OUTPUT_FILE)
        print(f"📂 Copie de {INPUT_FILE} → {OUTPUT_FILE}")
    else:
        print(f"📂 Chargement de {OUTPUT_FILE} (existant)")

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    # En-têtes nouvelles colonnes (seulement si vides)
    new_headers = {
        COL_ODOO:     ("STATUT ODOO CRM",                    25),
        COL_LINKEDIN: ("🔗 URL LINKEDIN\n(coller ici)",       35),
        COL_EMAIL:    ("📧 EMAIL\n(Zeliq)",                   30),
        COL_PHONE:    ("📞 TÉLÉPHONE\n(Zeliq)",               18),
        COL_ZELIQ_ST: ("⚡ STATUT\nENRICHISSEMENT ZELIQ",     24),
    }
    for col, (label, width) in new_headers.items():
        c = ws.cell(row=1, column=col)
        if not c.value:
            c.value     = label
            c.font      = Font(name='Arial', bold=True, size=9, color="FFFFFF")
            c.fill      = fill(BG["header"])
            c.alignment = align('center')
            c.border    = border()
            ws.column_dimensions[c.column_letter].width = width
    ws.row_dimensions[1].height = 40

    wb.save(OUTPUT_FILE)
    print(f"✅ Fichier prêt : {OUTPUT_FILE}\n")
    return wb, ws


# ══════════════════════════════════════════════════════════════════════════════
#  ODOO
# ══════════════════════════════════════════════════════════════════════════════

def odoo_connect():
    print(f"{'─'*55}\n  Connexion Odoo…")
    common = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/common")
    uid    = common.authenticate(ODOO_DB, ODOO_USER, ODOO_API_KEY, {})
    if not uid:
        print("❌ Authentification Odoo échouée."); sys.exit(1)
    models = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/object")
    print(f"  ✅ Connecté (UID={uid})\n")
    return models, uid


def odoo_clean_name(raw):
    name = str(raw or '').split('\n')[0].strip()
    name = re.sub(r'\s*\(.*?\)', '', name).strip()
    for s in [" SAS"," SA"," SARL"," SASU"," SNC"," INC."," INC"," LLC"," GROUP"," GROUPE"]:
        if name.upper().endswith(s): name = name[:-len(s)].strip()
    return name


def odoo_search(models, uid, company_name):
    clean = odoo_clean_name(company_name)
    if not clean: return []
    try:
        domain = ['|','|',
                  ['partner_name','ilike',clean],
                  ['partner_id.name','ilike',clean],
                  ['name','ilike',clean]]
        ids = models.execute_kw(ODOO_DB, uid, ODOO_API_KEY,
                                'crm.lead','search',[domain],{'limit':10})
        if not ids: return []
        return models.execute_kw(ODOO_DB, uid, ODOO_API_KEY,
                                 'crm.lead','read',[ids],
                                 {'fields':['name','partner_name','stage_id',
                                            'probability','user_id','active']})
    except Exception as e:
        print(f"   ⚠️  Erreur Odoo : {e}"); return []


def odoo_format(leads):
    if not leads: return "⚪ Aucune opportunité", "gray"
    lines, best = [], "gray"
    for l in sorted(leads, key=lambda x: x.get('active',False), reverse=True)[:3]:
        stage  = (l.get('stage_id') or [None,'?'])[1]
        proba  = l.get('probability', 0)
        user   = (l.get('user_id')   or [None,'?'])[1]
        name   = l.get('name','?')
        active = l.get('active', True)
        if not active:   e,bg,st = "🔴","red",  "Perdue/Archivée"
        elif proba==100: e,bg,st = "🟢","green","Gagnée"
        else:            e,bg,st = "🔵","blue", f"En cours ({int(proba)}%)"
        lines.append(f"{e} {name}\n   {stage} | {user} | {st}")
        if bg=="green": best="green"
        elif bg=="blue" and best!="green": best="blue"
        elif bg=="red"  and best=="gray":  best="red"
    suffix = f"\n({len(leads)} opp.)" if len(leads)>1 else ""
    return "\n\n".join(lines)+suffix, best


def odoo_create_lead(models, uid, company, signal, accroche, dg, daf, email, phone):
    contact = (dg or daf or '').split('\n')[0].strip()
    vals = {
        'name':         f"Prospection Assurance Transport — {company}",
        'partner_name': company,
        'description':  f"SIGNAL :\n{signal}\n\nACCROCHE :\n{accroche}\n\nCONTACT :\n{contact}",
        'type':         'opportunity',
    }
    if email and '@' in str(email): vals['email_from'] = str(email)
    if phone and len(str(phone)) > 5: vals['phone'] = str(phone)
    try:
        lead_id = models.execute_kw(ODOO_DB, uid, ODOO_API_KEY,
                                    'crm.lead','create',[vals])
        return lead_id, None
    except Exception as e:
        return None, str(e)


# ══════════════════════════════════════════════════════════════════════════════
#  ZELIQ
# ══════════════════════════════════════════════════════════════════════════════

def parse_contact_name(raw):
    """Extrait prénom/nom depuis une cellule décisionnaire."""
    if not raw: return '', ''
    line = str(raw).split('\n')[0].strip()
    for t in ['PDG','DG','DAF','CFO','CEO','CTO','Président','Directeur',
              'Dir.','Prés.','Fondateur','Chef']:
        line = re.sub(rf'\b{t}\b','', line, flags=re.IGNORECASE)
    line = re.sub(r'[–\-\(].*','', line).strip()
    parts = line.split()
    if len(parts) >= 2: return parts[0], ' '.join(parts[1:])
    return line, ''


def zeliq_enrich_row(row_data):
    """
    Enrichit téléphone + email pour une ligne.
    Retourne (phone, email, status_text, status_bg)
    """
    linkedin = row_data.get('linkedin','').strip()
    first    = row_data.get('first','')
    last     = row_data.get('last','')
    company  = row_data.get('company','')

    has_linkedin = bool(linkedin and 'linkedin.com' in linkedin)
    has_name     = bool(first and last and company)

    if not has_linkedin and not has_name:
        return '', '', "⚠️ Données insuffisantes\n(coller URL LinkedIn col. P)", "yellow"

    phone, email = '', ''

    # ── Téléphone ─────────────────────────────────────────────────────────────
    payload_p = {"callback_url": "https://webhook.site/zeliq-gsaprado"}
    if has_linkedin: payload_p["linkedin_url"] = linkedin
    try:
        r = requests.post(f"{ZELIQ_BASE}/contact/enrich/phone",
                          headers=ZELIQ_HDR, json=payload_p, timeout=30)
        if r.status_code == 200:
            phone = r.json().get('contact',{}).get('most_probable_phone','')
    except Exception as e:
        print(f"     ⚠️  Zeliq phone error: {e}")

    # ── Email ─────────────────────────────────────────────────────────────────
    payload_e = {"callback_url": "https://webhook.site/zeliq-gsaprado"}
    if has_linkedin: payload_e["linkedin_url"] = linkedin
    if has_name:
        payload_e.update({"first_name": first, "last_name": last, "company": company})
    try:
        r = requests.post(f"{ZELIQ_BASE}/contact/enrich/email",
                          headers=ZELIQ_HDR, json=payload_e, timeout=30)
        if r.status_code == 200:
            email = r.json().get('contact',{}).get('most_probable_email','')
    except Exception as e:
        print(f"     ⚠️  Zeliq email error: {e}")

    # ── Statut ────────────────────────────────────────────────────────────────
    if phone and email:  st, bg = "✅ Complet",         "green"
    elif phone or email: st, bg = "⚠️ Partiel",         "yellow"
    else:                st, bg = "❌ Non trouvé",       "red"

    return phone, email, st, bg


# ══════════════════════════════════════════════════════════════════════════════
#  ACTIONS
# ══════════════════════════════════════════════════════════════════════════════

def action_check_odoo(ws, models, uid):
    print(f"\n{'─'*55}")
    print("  1. VÉRIFICATION ODOO CRM")
    print(f"{'─'*55}")
    count = 0
    for row in range(FIRST_ROW, ws.max_row + 1):
        raw = ws.cell(row=row, column=COL_ENTREPRISE).value
        if not raw: continue
        name = str(raw).split('\n')[0].strip()
        print(f"  [{row-1:02d}] {name[:35]:<35}", end="… ", flush=True)
        leads         = odoo_search(models, uid, name)
        text, bg      = odoo_format(leads)
        write(ws, row, COL_ODOO, text, bg=bg)
        ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 70, 70)
        if not leads:               print("⚪ Aucune")
        elif bg == "green":         print(f"🟢 Gagnée ({len(leads)})")
        elif bg == "blue":          print(f"🔵 En cours ({len(leads)})")
        else:                       print(f"🔴 Archivée ({len(leads)})")
        count += 1
    print(f"\n  ✅ {count} entreprises vérifiées.")


def action_zeliq_enrich(ws):
    print(f"\n{'─'*55}")
    print("  2. ENRICHISSEMENT ZELIQ — compte par compte")
    print(f"{'─'*55}")
    print("""  Pour chaque contact vous pouvez :
    o   → Enrichir (téléphone + email)
    l   → Coller une URL LinkedIn puis enrichir
    s   → Passer (skip)
    q   → Quitter et sauvegarder
  """)

    # Recharger le fichier pour avoir les dernières saisies
    wb2 = load_workbook(OUTPUT_FILE)
    ws2 = wb2.active

    enriched = skipped = 0

    for row in range(FIRST_ROW, ws2.max_row + 1):
        raw = ws2.cell(row=row, column=COL_ENTREPRISE).value
        if not raw: continue

        company  = odoo_clean_name(raw)
        linkedin = str(ws2.cell(row=row, column=COL_LINKEDIN).value or '').strip()
        dg_raw   = ws2.cell(row=row, column=COL_DG).value
        daf_raw  = ws2.cell(row=row, column=COL_DAF).value
        first, last = parse_contact_name(dg_raw or daf_raw)

        cur_phone = str(ws2.cell(row=row, column=COL_PHONE).value or '').strip()
        cur_email = str(ws2.cell(row=row, column=COL_EMAIL).value or '').strip()
        already   = (cur_phone and cur_email
                     and "Non" not in cur_phone and "Non" not in cur_email)

        # ── Affichage fiche ───────────────────────────────────────────────────
        print(f"\n  {'═'*50}")
        print(f"  [{row-1:02d}] {company}")
        print(f"  {'─'*50}")
        contact_disp = str(dg_raw or daf_raw or "Non identifié").split("\n")[0].strip()
        print(f"  Contact     : {contact_disp}")
        print(f"  LinkedIn    : {linkedin if linkedin else '(non renseigné)'}")
        print(f"  Tél. actuel : {cur_phone if cur_phone else '—'}")
        print(f"  Email actuel: {cur_email if cur_email else '—'}")
        if already:
            print("  ✅ Déjà enrichi — vous pouvez relancer avec [o] pour rafraîchir")

        # ── Choix utilisateur ─────────────────────────────────────────────────
        rep = input("  → [o]Enrichir  [l]LinkedIn  [s]Skip  [q]Quitter : ").strip().lower()

        if rep == 'q':
            print("  ↩  Arrêt demandé.")
            break

        elif rep == 's':
            print("  ↩  Ignoré.")
            skipped += 1
            continue

        elif rep == 'l':
            url_input = input("  → Collez l'URL LinkedIn : ").strip()
            if 'linkedin.com' in url_input:
                linkedin = url_input
                write(ws2, row, COL_LINKEDIN, linkedin, bg="linkedin")
                wb2.save(OUTPUT_FILE)
                print(f"  ✅ URL enregistrée.")
            else:
                print("  ⚠️  URL invalide — on tente quand même avec le nom.")
            rep = 'o'

        if rep == 'o':
            has_linkedin = bool(linkedin and 'linkedin.com' in linkedin)
            has_name     = bool(first and last and company)

            if not has_linkedin and not has_name:
                print("  ⚠️  Pas d'URL LinkedIn ni de nom identifié.")
                print("       Utilisez [l] pour coller une URL LinkedIn.")
                write(ws2, row, COL_ZELIQ_ST,
                      "⚠️ URL LinkedIn manquante\n(utiliser option [l])", bg="yellow")
                wb2.save(OUTPUT_FILE)
                skipped += 1
                continue

            print("  ⚡ Enrichissement en cours…", end=" ", flush=True)

            phone, email, status, bg = zeliq_enrich_row({
                'linkedin': linkedin,
                'first': first, 'last': last,
                'company': company,
            })

            write(ws2, row, COL_PHONE,
                  phone or "Non trouvé",
                  bg="green" if phone else "red", bold=bool(phone))
            write(ws2, row, COL_EMAIL,
                  email or "Non trouvé",
                  bg="green" if email else "red", bold=bool(email))
            write(ws2, row, COL_ZELIQ_ST, status, bg=bg)
            ws2.row_dimensions[row].height = max(
                ws2.row_dimensions[row].height or 70, 70)

            print(f"📞 {phone or 'Non trouvé'}  |  📧 {email or 'Non trouvé'}")
            enriched += 1
            wb2.save(OUTPUT_FILE)   # sauvegarde après chaque enrichissement
            time.sleep(0.5)

    wb2.save(OUTPUT_FILE)
    print(f"\n  ✅ {enriched} enrichis, {skipped} ignorés → {OUTPUT_FILE}")


def action_create_leads(ws, models, uid):
    print(f"\n{'─'*55}")
    print("  3. CRÉATION LEADS ODOO")
    print(f"{'─'*55}")
    print("  Seules les lignes '⚪ Aucune opportunité' seront proposées.\n")

    # Recharger pour avoir les données Zeliq à jour
    wb2 = load_workbook(OUTPUT_FILE)
    ws2 = wb2.active

    created = skipped = 0
    for row in range(FIRST_ROW, ws2.max_row + 1):
        raw = ws2.cell(row=row, column=COL_ENTREPRISE).value
        if not raw: continue

        odoo_val = str(ws2.cell(row=row, column=COL_ODOO).value or '')
        if '⚪' not in odoo_val:
            continue  # déjà une opportunité

        company  = odoo_clean_name(raw)
        dg       = str(ws2.cell(row=row, column=COL_DG).value  or '')
        daf      = str(ws2.cell(row=row, column=COL_DAF).value or '')
        signal   = str(ws2.cell(row=row, column=COL_SIGNAL).value   or '')
        accroche = str(ws2.cell(row=row, column=COL_ACCROCHE).value or '')
        email    = str(ws2.cell(row=row, column=COL_EMAIL).value    or '')
        phone    = str(ws2.cell(row=row, column=COL_PHONE).value    or '')

        contact = (dg or daf or '').split('\n')[0].strip()
        print(f"  [{row-1:02d}] {company}")
        print(f"       Contact : {contact or 'non identifié'}")
        print(f"       Signal  : {signal[:70]}")
        if email and '@' in email: print(f"       Email   : {email}")
        if phone and len(phone)>5: print(f"       Tél.    : {phone}")

        rep = input("\n       → Créer ce lead dans Odoo ? [o/N] : ").strip().lower()
        if rep not in ('o','oui','y','yes'):
            print("       ↩  Ignoré.\n")
            skipped += 1
            continue

        lead_id, err = odoo_create_lead(
            models, uid, company, signal, accroche, dg, daf,
            email if '@' in email else '',
            phone if len(phone)>5 else ''
        )
        if lead_id:
            url = f"{ODOO_URL}/odoo/crm/{lead_id}"
            write(ws2, row, COL_ODOO,
                  f"🆕 Lead créé (ID:{lead_id})\n{url}",
                  bg="green", bold=True)
            print(f"       ✅ Lead créé → {url}\n")
            created += 1
        else:
            print(f"       ❌ Erreur Odoo : {err}\n")

    wb2.save(OUTPUT_FILE)
    print(f"  ✅ {created} leads créés, {skipped} ignorés.")


# ══════════════════════════════════════════════════════════════════════════════
#  MENU
# ══════════════════════════════════════════════════════════════════════════════

def menu(ws, models, uid):
    while True:
        print(f"""
{'═'*55}
  AGENT PROSPECTION GSA PRADO
  Source  : {INPUT_FILE}
  Sortie  : {OUTPUT_FILE}
{'═'*55}
  1  →  Vérifier les opportunités Odoo
  2  →  Enrichir contacts via Zeliq
         (téléphone + email, déclenché manuellement)
  3  →  Créer les leads manquants dans Odoo
         (confirmation ligne par ligne)
  4  →  Tout faire (1 puis 2 puis 3)
  0  →  Quitter
{'─'*55}""")
        choice = input("  Votre choix : ").strip()

        if choice == '0':
            print("\n  👋 Au revoir !\n")
            break

        elif choice == '1':
            action_check_odoo(ws, models, uid)
            wb_tmp = load_workbook(OUTPUT_FILE)
            wb_tmp.active = wb_tmp.active  # keep active
            # re-save via ws
            ws.parent.save(OUTPUT_FILE)
            print(f"\n  💾 Sauvegardé → {OUTPUT_FILE}")

        elif choice == '2':
            action_zeliq_enrich(ws)
            # ws rechargé dans la fonction

        elif choice == '3':
            action_create_leads(ws, models, uid)

        elif choice == '4':
            action_check_odoo(ws, models, uid)
            ws.parent.save(OUTPUT_FILE)
            action_zeliq_enrich(ws)
            # Recharger ws après Zeliq
            wb_new = load_workbook(OUTPUT_FILE)
            ws_new = wb_new.active
            action_create_leads(ws_new, models, uid)
            print(f"\n  💾 Sauvegardé final → {OUTPUT_FILE}")

        else:
            print("  ❓ Choix invalide, réessayez.")

        input("\n  Appuyez sur Entrée pour revenir au menu… ")


# ══════════════════════════════════════════════════════════════════════════════
#  POINT D'ENTRÉE
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print(f"""
{'═'*55}
  AGENT PROSPECTION GSA PRADO — v3
  {datetime.now().strftime('%d/%m/%Y %H:%M')}
{'═'*55}""")

    wb, ws = init_file()
    models, uid = odoo_connect()
    menu(ws, models, uid)
