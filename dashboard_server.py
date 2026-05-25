"""
=============================================================
  AGENT PROSPECTION GSA PRADO — Dashboard Web Local
  
  Lance un serveur local qui expose une API REST + interface web
  Lecture du fichier Excel + actions Zeliq + Odoo depuis le navigateur

  USAGE :
    pip install flask flask-cors requests openpyxl
    python dashboard_server.py
    → Ouvrir http://localhost:5000 dans Chrome
=============================================================
"""

import xmlrpc.client, re, os, sys, time, json, threading, shutil
from datetime import datetime

try:
    from flask import Flask, jsonify, request, send_from_directory
    from flask_cors import CORS
except ImportError:
    print("❌ Lancez : pip install flask flask-cors requests openpyxl"); sys.exit(1)

try:
    import requests as req_lib
except ImportError:
    print("❌ Lancez : pip install requests"); sys.exit(1)

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("❌ Lancez : pip install openpyxl"); sys.exit(1)

# ══════════════════════════════════════════════════════════════════════════════
#  CONFIG
# ══════════════════════════════════════════════════════════════════════════════

ODOO_URL     = "https://gsa-prado.odoo.com"
ODOO_DB      = "unikerp-gsaprado-prod-13772120"
ODOO_USER    = "georges-eric.michel@gsaprado.fr"
ODOO_API_KEY = "b4b43e5d24ac0631710e427dbfacc7951bcc7095"

ZELIQ_API_KEY = "sk-c4ce11eee8cfc5bf38c6256bbb27dd0dc85e65c68a52f740"
ZELIQ_BASE    = "https://api.zeliq.com/api"
ZELIQ_HDR     = {"x-api-key": ZELIQ_API_KEY, "Content-Type": "application/json"}

INPUT_FILE  = "prospection_v2.xlsx"
OUTPUT_FILE = "prospection_v3_final.xlsx"
PORT        = 5000

# Colonnes
COL = dict(
    ENTREPRISE=1, SECTEUR=2, VILLE=3, CA=4, RESULTAT=5, EVO=6, EFFECTIF=7,
    DG=8, DAF=9, SC=10, SIGNAL=11, ACCROCHE=12, ACTU=13, PRIORITE=14,
    ODOO=15, LINKEDIN=16, EMAIL=17, PHONE=18, ZELIQ_ST=19
)
FIRST_ROW = 2

# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def fill(c):    return PatternFill("solid", fgColor=c)
def bdr():
    s = Side(style='thin', color="BDC3C7")
    return Border(left=s, right=s, top=s, bottom=s)

BG = dict(
    header="1A3A5C", green="D5F5E3", blue="D6EAF8", red="FADBD8",
    gray="F2F3F4", yellow="FEF9E7", linkedin="E8F4FD", white="FFFFFF"
)

def write_cell(ws, row, col, value, bg="white", bold=False):
    c = ws.cell(row=row, column=col, value=str(value) if value else '')
    c.font      = Font(name='Arial', size=9, bold=bold, color="1A1A2E")
    c.fill      = fill(BG.get(bg, bg))
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    c.border    = bdr()
    ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 70, 70)
    return c

def clean_name(raw):
    name = str(raw or '').split('\n')[0].strip()
    name = re.sub(r'\s*\(.*?\)', '', name).strip()
    for s in [" SAS"," SA"," SARL"," SASU"," SNC"," INC."," INC"," LLC"," GROUP"," GROUPE"]:
        if name.upper().endswith(s): name = name[:-len(s)].strip()
    return name

def parse_contact(raw):
    if not raw: return '', ''
    line = str(raw).split('\n')[0].strip()
    for t in ['PDG','DG','DAF','CFO','CEO','Président','Directeur','Dir.','Fondateur']:
        line = re.sub(rf'\b{t}\b', '', line, flags=re.IGNORECASE)
    line = re.sub(r'[–\-\(].*', '', line).strip()
    parts = line.split()
    return (parts[0], ' '.join(parts[1:])) if len(parts) >= 2 else (line, '')

def load_output():
    if not os.path.exists(OUTPUT_FILE):
        if os.path.exists(INPUT_FILE):
            shutil.copy(INPUT_FILE, OUTPUT_FILE)
        else:
            return None, None
    wb = load_workbook(OUTPUT_FILE)
    return wb, wb.active

def cell_val(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v else ''

# ══════════════════════════════════════════════════════════════════════════════
#  ODOO
# ══════════════════════════════════════════════════════════════════════════════

_odoo_models = None
_odoo_uid    = None

def odoo_connect():
    global _odoo_models, _odoo_uid
    common = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/common")
    _odoo_uid    = common.authenticate(ODOO_DB, ODOO_USER, ODOO_API_KEY, {})
    _odoo_models = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/object")
    return bool(_odoo_uid)

def odoo_search(company):
    clean = clean_name(company)
    if not clean or not _odoo_uid: return []
    try:
        domain = ['|','|',
                  ['partner_name','ilike',clean],
                  ['partner_id.name','ilike',clean],
                  ['name','ilike',clean]]
        ids = _odoo_models.execute_kw(ODOO_DB, _odoo_uid, ODOO_API_KEY,
                                      'crm.lead','search',[domain],{'limit':5})
        if not ids: return []
        leads = _odoo_models.execute_kw(ODOO_DB, _odoo_uid, ODOO_API_KEY,
                                        'crm.lead','read',[ids],
                                        {'fields':['name','stage_id','probability',
                                                   'user_id','active']})
        return leads
    except: return []

def odoo_create(company, signal, accroche, dg, daf, email, phone):
    contact = (dg or daf or '').split('\n')[0].strip()
    vals = {
        'name':         f"Prospection Assurance Transport — {company}",
        'partner_name': company,
        'description':  f"SIGNAL :\n{signal}\n\nACCROCHE :\n{accroche}\n\nCONTACT :\n{contact}",
        'type':         'opportunity',
    }
    if email and '@' in str(email): vals['email_from'] = str(email)
    if phone and len(str(phone)) > 5: vals['phone'] = str(phone)
    lead_id = _odoo_models.execute_kw(ODOO_DB, _odoo_uid, ODOO_API_KEY,
                                      'crm.lead','create',[vals])
    return lead_id

# ══════════════════════════════════════════════════════════════════════════════
#  ZELIQ
# ══════════════════════════════════════════════════════════════════════════════

def zeliq_enrich(linkedin, first, last, company):
    phone, email = '', ''
    has_li   = bool(linkedin and 'linkedin.com' in linkedin)
    has_name = bool(first and last and company)
    if not has_li and not has_name:
        return '', '', 'missing'

    payload_p = {"callback_url": "https://webhook.site/zeliq-gsaprado"}
    if has_li: payload_p["linkedin_url"] = linkedin
    try:
        r = req_lib.post(f"{ZELIQ_BASE}/contact/enrich/phone",
                         headers=ZELIQ_HDR, json=payload_p, timeout=30)
        if r.status_code == 200:
            phone = r.json().get('contact',{}).get('most_probable_phone','')
    except: pass

    payload_e = {"callback_url": "https://webhook.site/zeliq-gsaprado"}
    if has_li: payload_e["linkedin_url"] = linkedin
    if has_name: payload_e.update({"first_name":first,"last_name":last,"company":company})
    try:
        r = req_lib.post(f"{ZELIQ_BASE}/contact/enrich/email",
                         headers=ZELIQ_HDR, json=payload_e, timeout=30)
        if r.status_code == 200:
            email = r.json().get('contact',{}).get('most_probable_email','')
    except: pass

    status = 'ok' if (phone and email) else ('partial' if (phone or email) else 'not_found')
    return phone, email, status

# ══════════════════════════════════════════════════════════════════════════════
#  FLASK APP
# ══════════════════════════════════════════════════════════════════════════════

app = Flask(__name__, static_folder='dashboard_static')
CORS(app)

@app.route('/')
def index():
    return send_from_directory('dashboard_static', 'index.html')

# ── Lire toutes les lignes ────────────────────────────────────────────────────
@app.route('/api/rows')
def api_rows():
    wb, ws = load_output()
    if not ws: return jsonify({'error': f"Fichier {OUTPUT_FILE} introuvable"}), 404
    rows = []
    for row in range(FIRST_ROW, ws.max_row + 1):
        entreprise = cell_val(ws, row, COL['ENTREPRISE'])
        if not entreprise: continue
        rows.append({
            'row':       row,
            'entreprise': entreprise,
            'secteur':   cell_val(ws, row, COL['SECTEUR']),
            'ville':     cell_val(ws, row, COL['VILLE']),
            'ca':        cell_val(ws, row, COL['CA']),
            'priorite':  cell_val(ws, row, COL['PRIORITE']),
            'dg':        cell_val(ws, row, COL['DG']),
            'daf':       cell_val(ws, row, COL['DAF']),
            'signal':    cell_val(ws, row, COL['SIGNAL']),
            'accroche':  cell_val(ws, row, COL['ACCROCHE']),
            'actu':      cell_val(ws, row, COL['ACTU']),
            'linkedin':  cell_val(ws, row, COL['LINKEDIN']),
            'email':     cell_val(ws, row, COL['EMAIL']),
            'phone':     cell_val(ws, row, COL['PHONE']),
            'odoo':      cell_val(ws, row, COL['ODOO']),
            'zeliq_st':  cell_val(ws, row, COL['ZELIQ_ST']),
        })
    return jsonify(rows)

# ── Vérifier Odoo pour une ligne ─────────────────────────────────────────────
@app.route('/api/odoo/check/<int:row>', methods=['POST'])
def api_odoo_check(row):
    wb, ws = load_output()
    company = cell_val(ws, row, COL['ENTREPRISE'])
    leads   = odoo_search(company)
    if not leads:
        text, bg = "⚪ Aucune opportunité", "gray"
    else:
        parts = []
        bg = "gray"
        for l in leads[:3]:
            stage  = (l.get('stage_id') or [None,'?'])[1]
            proba  = l.get('probability', 0)
            user   = (l.get('user_id')   or [None,'?'])[1]
            active = l.get('active', True)
            if not active:    e,bg = "🔴","red"
            elif proba==100:  e,bg = "🟢","green"
            else:             e,bg = "🔵","blue"
            parts.append(f"{e} {l['name']} | {stage} | {user}")
        text = "\n".join(parts) + (f"\n({len(leads)} opp.)" if len(leads)>1 else "")
    write_cell(ws, row, COL['ODOO'], text, bg=bg)
    wb.save(OUTPUT_FILE)
    return jsonify({'status': bg, 'text': text})

# ── Créer lead Odoo ───────────────────────────────────────────────────────────
@app.route('/api/odoo/create/<int:row>', methods=['POST'])
def api_odoo_create(row):
    wb, ws = load_output()
    company  = cell_val(ws, row, COL['ENTREPRISE'])
    signal   = cell_val(ws, row, COL['SIGNAL'])
    accroche = cell_val(ws, row, COL['ACCROCHE'])
    dg       = cell_val(ws, row, COL['DG'])
    daf      = cell_val(ws, row, COL['DAF'])
    email    = cell_val(ws, row, COL['EMAIL'])
    phone    = cell_val(ws, row, COL['PHONE'])
    try:
        lead_id = odoo_create(clean_name(company), signal, accroche, dg, daf, email, phone)
        url = f"{ODOO_URL}/odoo/crm/{lead_id}"
        write_cell(ws, row, COL['ODOO'], f"🆕 Lead créé (ID:{lead_id})\n{url}", bg="green", bold=True)
        wb.save(OUTPUT_FILE)
        return jsonify({'lead_id': lead_id, 'url': url})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── Enrichir via Zeliq ────────────────────────────────────────────────────────
@app.route('/api/zeliq/enrich/<int:row>', methods=['POST'])
def api_zeliq_enrich(row):
    data = request.json or {}
    wb, ws = load_output()
    company  = clean_name(cell_val(ws, row, COL['ENTREPRISE']))
    linkedin = data.get('linkedin') or cell_val(ws, row, COL['LINKEDIN'])
    dg_raw   = cell_val(ws, row, COL['DG'])
    daf_raw  = cell_val(ws, row, COL['DAF'])
    first, last = parse_contact(dg_raw or daf_raw)

    if linkedin:
        write_cell(ws, row, COL['LINKEDIN'], linkedin, bg="linkedin")

    phone, email, status = zeliq_enrich(linkedin, first, last, company)

    if phone: write_cell(ws, row, COL['PHONE'], phone, bg="green", bold=True)
    else:     write_cell(ws, row, COL['PHONE'], "Non trouvé", bg="red")
    if email: write_cell(ws, row, COL['EMAIL'], email, bg="green", bold=True)
    else:     write_cell(ws, row, COL['EMAIL'], "Non trouvé", bg="red")

    status_labels = {'ok':'✅ Complet','partial':'⚠️ Partiel',
                     'not_found':'❌ Non trouvé','missing':'⚠️ Données insuffisantes'}
    status_bgs    = {'ok':'green','partial':'yellow','not_found':'red','missing':'yellow'}
    write_cell(ws, row, COL['ZELIQ_ST'], status_labels.get(status,'?'),
               bg=status_bgs.get(status,'gray'))
    wb.save(OUTPUT_FILE)
    return jsonify({'phone': phone, 'email': email, 'status': status})

# ── Sauvegarder URL LinkedIn ──────────────────────────────────────────────────
@app.route('/api/linkedin/<int:row>', methods=['POST'])
def api_save_linkedin(row):
    data = request.json or {}
    url  = data.get('linkedin','').strip()
    if not url: return jsonify({'error': 'URL vide'}), 400
    wb, ws = load_output()
    write_cell(ws, row, COL['LINKEDIN'], url, bg="linkedin")
    wb.save(OUTPUT_FILE)
    return jsonify({'saved': url})

if __name__ == '__main__':
    print(f"\n{'═'*55}")
    print(f"  AGENT PROSPECTION GSA PRADO — Dashboard")
    print(f"  {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print(f"{'═'*55}")

    # Init fichier
    if not os.path.exists(OUTPUT_FILE) and os.path.exists(INPUT_FILE):
        shutil.copy(INPUT_FILE, OUTPUT_FILE)
        print(f"  📂 {INPUT_FILE} → {OUTPUT_FILE}")
    elif not os.path.exists(INPUT_FILE):
        print(f"  ❌ Fichier introuvable : {INPUT_FILE}"); sys.exit(1)

    # Connexion Odoo
    print("  🔌 Connexion Odoo…", end=" ")
    if odoo_connect(): print("✅")
    else: print("❌ Vérifiez vos identifiants")

    # Dossier static
    os.makedirs('dashboard_static', exist_ok=True)

    print(f"\n  🌐 Dashboard disponible sur : http://localhost:{PORT}")
    print(f"  ⌨️  Ctrl+C pour arrêter\n{'═'*55}\n")
    app.run(host='0.0.0.0', port=PORT, debug=False)
