"""
Ajoute les nouvelles entreprises du Google Sheets dans le fichier Excel
et vérifie pour chacune s'il existe une opportunité dans Odoo.

Usage : python check_new_companies.py
"""
import xmlrpc.client, re, os, sys, shutil
from datetime import datetime

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

# ── Config ────────────────────────────────────────────────────────────────────
ODOO_URL     = "https://gsa-prado.odoo.com"
ODOO_DB      = "unikerp-gsaprado-prod-13772120"
ODOO_USER    = "georges-eric.michel@gsaprado.fr"
ODOO_API_KEY = "b4b43e5d24ac0631710e427dbfacc7951bcc7095"

INPUT_FILE  = "prospection_v3_final.xlsx" if os.path.exists("prospection_v3_final.xlsx") \
              else "prospection_v2.xlsx"
OUTPUT_FILE = "prospection_v3_final.xlsx"

COL_ENTREPRISE = 1
COL_SECTEUR    = 2
COL_VILLE      = 3
COL_ODOO       = 15
FIRST_ROW      = 2

# ── Nouvelles entreprises à ajouter ──────────────────────────────────────────
# (nom, secteur, ville) — issues du Google Sheets
NOUVELLES = [
    ("EPC FRANCE",                          "Fabrication explosifs & forage-minage",         "Saint-Martin-de-Crau (13)"),
    ("ARKADIA Group",                       "Ingénierie & conseil – énergie, aéronautique",   "Aix-en-Provence (13)"),
    ("SESSÙN",                              "Mode féminine – retail & export",                "Marseille (13)"),
    ("Altereo",                             "Ingénierie environnementale & eau",              "Venelles (13)"),
    ("CHU de Nice / EMRC ONCOPACA",         "Hôpital public – santé",                        "Aix-en-Provence (13)"),
    ("LE TEMPS DES CERISES JEANS",          "Mode – jeans & prêt-à-porter",                  "Marseille (13)"),
    ("PROFROID",                            "Fabrication équipements réfrigération",          "Aubagne (13)"),
    ("Provepharm",                          "Pharmaceutique – produits hospitaliers",         "Marseille (13)"),
    ("JEFCO",                               "Fabrication peintures professionnelles",         "Marseille (13)"),
    ("Laboratoire d'Astrophysique Marseille","Recherche scientifique (CNRS/AMU)",             "Marseille (13)"),
    ("CEREG",                               "Bureau d'ingénierie – aménagement territoire",  "Montpellier (34)"),
    ("LPB Les Petites Bombes",              "Prêt-à-porter féminin",                         "La Ciotat (13)"),
    ("EDL",                                 "IT médical – imagerie médicale",                "Berre-l'Étang (13)"),
    ("Olea Medical",                        "Software imagerie médicale (IRM/scanner)",       "La Ciotat (13)"),
    ("ECF PRO",                             "Formation professionnelle – transport & BTP",    "Marseille (13)"),
    ("Christine Laure",                     "Mode féminine – retail",                        "Marseille (13)"),
    ("PharmaBest",                          "Réseau pharmacies – 135 officines France",       "Marseille (13)"),
    ("JOTT",                                "Mode – doudounes urbaines & retail",             "Marseille (13)"),
    ("Pharma & Beauty Group",               "Sous-traitance cosmétique & pharmaceutique",    "Saint-Chamas (13)"),
    ("KAPORAL",                             "Mode – jeans & streetwear",                     "Marseille (13)"),
    ("Enovacom",                            "Logiciels santé – interopérabilité données",    "Marseille (13)"),
    ("RDT13",                               "Transport en commun & fret ferroviaire",         "Aix-en-Provence (13)"),
]

# ── Style helpers ─────────────────────────────────────────────────────────────
BG = dict(header="1A3A5C", green="D5F5E3", blue="D6EAF8",
          red="FADBD8", gray="F2F3F4", white="FFFFFF",
          light_blue="D6E4F0", new_row="EEF6FF")

def fill(c): return PatternFill("solid", fgColor=BG.get(c, c))
def bdr():
    s = Side(style='thin', color="BDC3C7")
    return Border(left=s, right=s, top=s, bottom=s)

def write(ws, row, col, value, bg="white", bold=False, color="1A1A2E"):
    c = ws.cell(row=row, column=col, value=str(value) if value else '')
    c.font      = Font(name='Arial', size=9, bold=bold, color=color)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    c.border    = bdr()
    ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 55, 55)
    return c

def clean(raw):
    name = str(raw or '').split('\n')[0].strip()
    name = re.sub(r'\s*\(.*?\)', '', name).strip()
    for s in [" SAS"," SA"," SARL"," SASU"," SNC"," INC."," INC"," LLC"]:
        if name.upper().endswith(s): name = name[:-len(s)].strip()
    return name.upper()

# ── Odoo ──────────────────────────────────────────────────────────────────────
def odoo_connect():
    common = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/common")
    uid    = common.authenticate(ODOO_DB, ODOO_USER, ODOO_API_KEY, {})
    if not uid: print("❌ Odoo auth failed"); sys.exit(1)
    models = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/object")
    print(f"✅ Odoo connecté (UID={uid})\n")
    return models, uid

def odoo_search(models, uid, name):
    c = clean(name)
    if not c: return []
    try:
        domain = ['|','|',
                  ['partner_name','ilike',c],
                  ['partner_id.name','ilike',c],
                  ['name','ilike',c]]
        ids = models.execute_kw(ODOO_DB, uid, ODOO_API_KEY,
                                'crm.lead','search',[domain],{'limit':5})
        if not ids: return []
        return models.execute_kw(ODOO_DB, uid, ODOO_API_KEY,
                                 'crm.lead','read',[ids],
                                 {'fields':['name','stage_id','probability','user_id','active']})
    except Exception as e:
        print(f"  ⚠️  {e}"); return []

def format_status(leads):
    if not leads: return "⚪ Aucune opportunité", "gray"
    lines, best = [], "gray"
    for l in leads[:3]:
        stage  = (l.get('stage_id') or [None,'?'])[1]
        proba  = l.get('probability', 0)
        user   = (l.get('user_id')   or [None,'?'])[1]
        active = l.get('active', True)
        if not active:     e, bg = "🔴", "red"
        elif proba == 100: e, bg = "🟢", "green"
        else:              e, bg = "🔵", "blue"
        lines.append(f"{e} {l['name']}\n   {stage} | {user}")
        if bg == "green": best = "green"
        elif bg == "blue" and best != "green": best = "blue"
        elif bg == "red"  and best == "gray":  best = "red"
    return "\n\n".join(lines) + (f"\n({len(leads)} opp.)" if len(leads)>1 else ""), best

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print(f"\n{'═'*55}")
    print(f"  Ajout nouvelles entreprises + vérification Odoo")
    print(f"  {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print(f"{'═'*55}")

    # Charger le fichier
    if not os.path.exists(INPUT_FILE):
        print(f"❌ Fichier introuvable : {INPUT_FILE}"); sys.exit(1)
    if INPUT_FILE != OUTPUT_FILE:
        shutil.copy(INPUT_FILE, OUTPUT_FILE)
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    # Récupérer les noms déjà présents dans le fichier
    existing_names = set()
    for row in range(FIRST_ROW, ws.max_row + 1):
        val = ws.cell(row=row, column=COL_ENTREPRISE).value
        if val:
            existing_names.add(clean(str(val)))

    print(f"  📋 {len(existing_names)} entreprises déjà dans le fichier")
    print(f"  ➕ {len(NOUVELLES)} nouvelles à traiter\n")

    models, uid = odoo_connect()

    added = already = 0
    for nom, secteur, ville in NOUVELLES:
        nom_clean = clean(nom)

        # Vérifier si déjà présente
        if nom_clean in existing_names:
            print(f"  ↩  {nom[:45]:<45} déjà présent")
            already += 1
            continue

        # Ajouter la ligne à la fin du fichier
        new_row = ws.max_row + 1
        write(ws, new_row, COL_ENTREPRISE, nom,     bg="light_blue", bold=True, color="1A3A5C")
        write(ws, new_row, COL_SECTEUR,    secteur, bg="new_row")
        write(ws, new_row, COL_VILLE,      ville,   bg="new_row")

        # Remplir les colonnes vides avec N/C pour cohérence visuelle
        for col in range(4, COL_ODOO):
            if col not in (COL_ENTREPRISE, COL_SECTEUR, COL_VILLE):
                c = ws.cell(row=new_row, column=col)
                if not c.value:
                    c.fill   = fill("new_row")
                    c.border = bdr()

        # Vérifier Odoo
        print(f"  🆕 {nom[:45]:<45}", end="… ", flush=True)
        leads    = odoo_search(models, uid, nom)
        text, bg = format_status(leads)
        write(ws, new_row, COL_ODOO, text, bg=bg)

        if not leads:       print("⚪ Aucune")
        elif bg == "green": print(f"🟢 Gagnée ({len(leads)})")
        elif bg == "blue":  print(f"🔵 En cours ({len(leads)})")
        else:               print(f"🔴 Archivée ({len(leads)})")

        existing_names.add(nom_clean)
        added += 1

    wb.save(OUTPUT_FILE)
    print(f"\n{'─'*55}")
    print(f"  ✅ {added} nouvelles entreprises ajoutées et vérifiées")
    print(f"  ↩  {already} déjà présentes, ignorées")
    print(f"  💾 Fichier mis à jour : {OUTPUT_FILE}")
    print(f"{'═'*55}\n")

if __name__ == "__main__":
    main()
