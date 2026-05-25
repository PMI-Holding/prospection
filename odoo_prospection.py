"""
=============================================================
  AGENT PROSPECTION - Vérification Odoo CRM
  GSA Prado - Courtage Assurance Transport
=============================================================

USAGE :
  1. Installez les dépendances :
       pip install openpyxl

  2. Placez ce script dans le même dossier que votre fichier
     Excel de prospection (prospection_v2.xlsx)

  3. Lancez :
       python odoo_prospection.py

  4. Le script génère : prospection_v2_odoo.xlsx
     avec une colonne "STATUT ODOO" remplie pour chaque entreprise.

CONFIGURATION :
  Modifiez les variables dans la section CONFIG ci-dessous.
=============================================================
"""

import xmlrpc.client
import re
import os
import sys
from datetime import datetime

# ── Tentative d'import openpyxl ──────────────────────────────────────────────
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("❌ Module 'openpyxl' manquant.")
    print("   Installez-le avec : pip install openpyxl")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
#  CONFIG — À adapter si besoin
# ══════════════════════════════════════════════════════════════════════════════

ODOO_URL      = "https://gsa-prado.odoo.com"
ODOO_DB       = "unikerp-gsaprado-prod-13772120"
ODOO_USER     = "georges-eric.michel@gsaprado.fr"
ODOO_API_KEY  = "b4b43e5d24ac0631710e427dbfacc7951bcc7095"

# Fichier Excel source (dans le même dossier que ce script)
INPUT_FILE    = "prospection_v2.xlsx"
OUTPUT_FILE   = "prospection_v2_odoo.xlsx"

# Colonne du fichier Excel qui contient le nom des entreprises (A = 1)
COL_ENTREPRISE = 1

# Ligne à partir de laquelle commencent les données (après l'en-tête)
FIRST_DATA_ROW = 2

# Colonne où écrire le statut Odoo (O = 15)
COL_ODOO_STATUS = 15

# ══════════════════════════════════════════════════════════════════════════════


# ── Couleurs ─────────────────────────────────────────────────────────────────
COLORS = {
    "header_bg":  "1A3A5C",
    "header_fg":  "FFFFFF",
    "won":        "D5F5E3",   # vert clair  — opportunité gagnée
    "open":       "D6EAF8",   # bleu clair  — opportunité ouverte
    "lost":       "FADBD8",   # rouge clair — opportunité perdue
    "none":       "F8F9FA",   # gris clair  — aucune opportunité
    "odd":        "F2F3F4",
    "even":       "FFFFFF",
    "border":     "BDC3C7",
}

def make_fill(color):
    return PatternFill("solid", fgColor=color)

def make_border():
    s = Side(style='thin', color=COLORS["border"])
    return Border(left=s, right=s, top=s, bottom=s)


# ── Détection automatique du nom de base ─────────────────────────────────────
def find_database(common):
    """
    Odoo SaaS : le nom de la DB n'est pas toujours identique au sous-domaine.
    On essaie d'abord de lister les DB, puis on teste des variantes courantes.
    """
    # 1. Tentative de listing (souvent bloqué sur SaaS)
    try:
        db_svc = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/db")
        dbs = db_svc.list()
        if dbs:
            print(f"   Bases disponibles : {dbs}")
            return dbs[0]   # prend la première si une seule
    except Exception:
        pass

    # 2. Si ODOO_DB est défini manuellement, l'utiliser directement
    if ODOO_DB:
        return ODOO_DB

    # 3. Dériver des variantes depuis l'URL
    host = ODOO_URL.replace("https://", "").replace("http://", "").split(".")[0]
    candidates = [
        host,
        host.replace("-", "_"),
        host.replace("_", "-"),
        host + "-main",
        host + "-prod",
        host + "-master",
        host.split("-")[0],          # "gsa" si "gsa-prado"
        "-".join(host.split("-")[1:]), # "prado" si "gsa-prado"
    ]
    # Dédoublonner tout en gardant l'ordre
    seen = set()
    candidates = [c for c in candidates if c and not (c in seen or seen.add(c))]

    print(f"   Recherche automatique parmi : {candidates}")
    for db in candidates:
        try:
            uid = common.authenticate(db, ODOO_USER, ODOO_API_KEY, {})
            if uid:
                print(f"   ✅ Base trouvée : '{db}'")
                return db
        except Exception as e:
            if "does not exist" in str(e):
                continue
            # Autre erreur (réseau, etc.) → on arrête
            break

    return None


# ── Connexion Odoo ────────────────────────────────────────────────────────────
def connect_odoo():
    print(f"\n{'='*55}")
    print(f"  Connexion à {ODOO_URL}")
    print(f"{'='*55}")
    try:
        common = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/common")

        # ── Détection de la base ──────────────────────────────────────────────
        db = find_database(common)
        if not db:
            print("\n❌ Impossible de déterminer le nom de la base de données.")
            print("\n   👉 Solution : ouvrez Odoo dans votre navigateur, puis :")
            print("      1. Allez dans  Paramètres → Technique → Base de données")
            print("      2. Ou regardez l'URL après connexion :")
            print("         https://gsa-prado.odoo.com/web#action=... ")
            print("         Le nom est visible dans  Paramètres → À propos")
            print("\n   Puis relancez le script en renseignant ODOO_DB manuellement")
            print("   dans la section CONFIG (ligne ~45 du script).")
            sys.exit(1)

        # ── Authentification ──────────────────────────────────────────────────
        uid = common.authenticate(db, ODOO_USER, ODOO_API_KEY, {})
        if not uid:
            print("❌ Authentification échouée.")
            print("   → Vérifiez que la clé API est active dans Odoo :")
            print("     Paramètres → Mon profil → Sécurité du compte → Clés API")
            sys.exit(1)

        models = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/object")
        print(f"✅ Connecté — Base : '{db}' | UID : {uid}")

        # Stocker le nom de DB trouvé pour les appels suivants
        connect_odoo.db = db
        return models, uid

    except Exception as e:
        print(f"❌ Erreur de connexion : {e}")
        sys.exit(1)


# ── Nettoyage du nom d'entreprise ─────────────────────────────────────────────
def clean_name(raw):
    """Retire les sauts de ligne, parenthèses et suffixes courants."""
    name = str(raw or "").strip()
    name = name.split("\n")[0].strip()          # première ligne seulement
    # Retire les mentions entre parenthèses : "NAOS (Bioderma...)" → "NAOS"
    name = re.sub(r'\s*\(.*?\)', '', name).strip()
    # Retire suffixes juridiques courants
    for suffix in [" SAS", " SA", " SARL", " SNC", " SASU", " INC.", " INC",
                   " LLC", " GROUP", " GROUPE", " HOLDING", " HOLDINGS"]:
        if name.upper().endswith(suffix):
            name = name[:-len(suffix)].strip()
    return name


# ── Recherche dans Odoo ───────────────────────────────────────────────────────
def search_odoo_leads(models, uid, company_name):
    """
    Cherche des opportunités/leads CRM correspondant à company_name.
    Retourne une liste de dicts avec les champs clés.
    """
    clean = clean_name(company_name)
    if not clean:
        return []

    try:
        # Recherche sur le nom du partenaire OU le nom de l'opportunité
        domain = [
            '|', '|',
            ['partner_name', 'ilike', clean],
            ['partner_id.name', 'ilike', clean],
            ['name', 'ilike', clean],
        ]

        db = connect_odoo.db   # nom de base détecté à la connexion

        lead_ids = models.execute_kw(
            db, uid, ODOO_API_KEY,
            'crm.lead', 'search',
            [domain],
            {'limit': 10}
        )

        if not lead_ids:
            return []

        leads = models.execute_kw(
            db, uid, ODOO_API_KEY,
            'crm.lead', 'read',
            [lead_ids],
            {'fields': [
                'name',           # nom de l'opportunité
                'partner_name',   # nom société (si pas de partenaire)
                'partner_id',     # partenaire lié
                'stage_id',       # étape du pipeline
                'probability',    # probabilité de closing
                'user_id',        # commercial assigné
                'date_deadline',  # date de closing prévue
                'active',         # actif ou archivé
                'type',           # lead ou opportunity
            ]}
        )
        return leads

    except Exception as e:
        print(f"   ⚠️  Erreur recherche '{clean}' : {e}")
        return []


# ── Formatage du statut ───────────────────────────────────────────────────────
def format_status(leads):
    """
    Retourne (texte_statut, couleur_bg) depuis la liste de leads trouvés.
    """
    if not leads:
        return "⚪ Aucune opportunité", COLORS["none"]

    # Trier : opportunités actives en premier
    active = [l for l in leads if l.get('active')]
    archived = [l for l in leads if not l.get('active')]

    lines = []
    best_color = COLORS["none"]

    for lead in (active + archived)[:3]:   # max 3 résultats affichés
        stage   = lead.get('stage_id', [None, '?'])[1] if lead.get('stage_id') else '?'
        proba   = lead.get('probability', 0)
        user    = lead.get('user_id', [None, '?'])[1] if lead.get('user_id') else '?'
        opp_name = lead.get('name', '?')
        is_active = lead.get('active', True)

        # Détecter si gagnée ou perdue d'après la probabilité / archivage
        if not is_active:
            emoji = "🔴"
            color = COLORS["lost"]
            status_txt = "Perdue/Archivée"
        elif proba == 100:
            emoji = "🟢"
            color = COLORS["won"]
            status_txt = "Gagnée"
        else:
            emoji = "🔵"
            color = COLORS["open"]
            status_txt = f"En cours ({int(proba)}%)"

        lines.append(f"{emoji} {opp_name}\n   Étape : {stage}\n   Commercial : {user}\n   Statut : {status_txt}")

        # Priorité couleur : vert > bleu > rouge
        if color == COLORS["won"]:
            best_color = color
        elif color == COLORS["open"] and best_color != COLORS["won"]:
            best_color = color
        elif color == COLORS["lost"] and best_color == COLORS["none"]:
            best_color = color

    total = len(leads)
    suffix = f"\n   ({total} opportunité(s) trouvée(s))" if total > 1 else ""
    return "\n\n".join(lines) + suffix, best_color


# ── Traitement du fichier Excel ───────────────────────────────────────────────
def process_excel(models, uid):
    # Vérification fichier source
    if not os.path.exists(INPUT_FILE):
        print(f"\n❌ Fichier introuvable : {INPUT_FILE}")
        print(f"   Placez '{INPUT_FILE}' dans le même dossier que ce script.")
        sys.exit(1)

    print(f"\n📂 Chargement de : {INPUT_FILE}")
    wb = load_workbook(INPUT_FILE)
    ws = wb.active

    # ── En-tête colonne Odoo ──────────────────────────────────────────────────
    header_cell = ws.cell(row=1, column=COL_ODOO_STATUS)
    header_cell.value     = "STATUT ODOO CRM"
    header_cell.font      = Font(name='Arial', bold=True, size=9, color=COLORS["header_fg"])
    header_cell.fill      = make_fill(COLORS["header_bg"])
    header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_cell.border    = make_border()
    ws.column_dimensions[header_cell.column_letter].width = 40
    ws.row_dimensions[1].height = 40

    # ── Traitement ligne par ligne ────────────────────────────────────────────
    max_row = ws.max_row
    print(f"   {max_row - 1} entreprises à vérifier\n")

    for row in range(FIRST_DATA_ROW, max_row + 1):
        raw_name = ws.cell(row=row, column=COL_ENTREPRISE).value
        if not raw_name:
            continue

        clean = clean_name(str(raw_name))
        print(f"  [{row-1:02d}/{max_row-1}] Recherche : '{clean}'", end=" ... ", flush=True)

        leads = search_odoo_leads(models, uid, clean)
        status_text, status_color = format_status(leads)

        if not leads:
            print("⚪ Aucune")
        elif any(not l.get('active') for l in leads):
            print(f"🔴 {len(leads)} archivée(s)")
        elif any(l.get('probability') == 100 for l in leads):
            print(f"🟢 Gagnée(s)")
        else:
            print(f"🔵 {len(leads)} en cours")

        # Écriture dans la cellule
        cell = ws.cell(row=row, column=COL_ODOO_STATUS)
        cell.value     = status_text
        cell.fill      = make_fill(status_color)
        cell.font      = Font(name='Arial', size=8, color="1A1A2E")
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border    = make_border()
        ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 70, 70)

    # ── Sauvegarde ────────────────────────────────────────────────────────────
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Fichier généré : {OUTPUT_FILE}")
    print(f"   Horodatage     : {datetime.now().strftime('%d/%m/%Y %H:%M')}")


# ── Point d'entrée ────────────────────────────────────────────────────────────
if __name__ == "__main__":
    models, uid = connect_odoo()
    process_excel(models, uid)
    print(f"\n{'='*55}")
    print("  Terminé ! Ouvrez le fichier Excel généré.")
    print(f"{'='*55}\n")
